import openpyxl
from openpyxl.utils import get_column_letter

def restore_and_round_values():
    try:
        # Загружаем оба файла
        input_wb = openpyxl.load_workbook('input.xlsx')
        output_wb = openpyxl.load_workbook('output.xlsx')
        
        input_sheet = input_wb.active
        output_sheet = output_wb.active
        
        print("Файлы успешно загружены")

        # Создаем новый файл для результатов
        final_wb = openpyxl.Workbook()
        final_sheet = final_wb.active
        
        # Копируем заголовки (если есть)
        for row in input_sheet.iter_rows(max_row=1):
            for cell in row:
                final_sheet[cell.coordinate].value = cell.value

        # Обрабатываем каждую строку
        for row in range(2, output_sheet.max_row + 1):
            # Проверяем, все ли нули в колонках G-M в output
            all_zeros = True
            for col in range(7, 14):  # Колонки G-M (7-13)
                if output_sheet.cell(row=row, column=col).value != 0:
                    all_zeros = False
                    break
            
            # Копируем данные
            for col in range(1, output_sheet.max_column + 1):
                if all_zeros and 7 <= col <= 13:  # Если нули в G-M, берем из input
                    value = input_sheet.cell(row=row, column=col).value
                else:  # Иначе берем из output
                    value = output_sheet.cell(row=row, column=col).value
                
                # Округляем значение в колонке G
                if col == 7 and isinstance(value, (int, float)):
                    value = round(float(value), 1)
                
                final_sheet.cell(row=row, column=col, value=value)

            # Выводим информацию о строке
            if all_zeros:
                print(f"Строка {row}: данные восстановлены из input.xlsx")
            else:
                print(f"Строка {row}: данные сохранены из output.xlsx")

        # Сохраняем результат
        final_wb.save('final_output.xlsx')
        print("Обработка завершена. Результат сохранен в final_output.xlsx")
        
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")

if __name__ == "__main__":
    restore_and_round_values()