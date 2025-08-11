import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException

def main():
    # Настройка Selenium WebDriver
    chrome_options = Options()
#    chrome_options.add_argument("--headless")  # Работа в фоновом режиме
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    
    # Укажите путь к вашему chromedriver
    service = Service('chromedriver.exe')
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    # Пауза 30 секунд после открытия браузера
    print("Ожидание 30 секунд после открытия браузера...")
    time.sleep(30)
    
    # Загрузка Excel файла
    try:
        wb = load_workbook('index.xlsx')
        ws = wb.active
        print("Excel файл успешно загружен")
    except Exception as e:
        print(f"Ошибка при загрузке Excel файла: {e}")
        driver.quit()
        return
    
    # Проход по строкам таблицы
    for row in range(1, ws.max_row + 1):
        link_cell = ws.cell(row=row, column=14)  # Колонка N
        link = link_cell.value
        
        # Пропуск строки без ссылки
        if not link or not isinstance(link, str) or not link.startswith(('http://', 'https://')):
            print(f"Строка {row}: пропущена (нет ссылки)")
            continue
        
        try:
            # Открытие страницы
            driver.get(link)
            time.sleep(1)  # Ожидание прогрузки страницы
            
            # Поиск и извлечение данных
            try:
                # Значение для колонки O
                span_o = driver.find_element(By.CSS_SELECTOR, 'span.ds-text.ds-text_weight_bold.ds-text_proportional.ds-text_typography_headline-1.ds-text_headline-1_tight.ds-text_headline-1_bold')
                ws.cell(row=row, column=15, value=span_o.text)  # Колонка O
                
                # Значение для колонки P (только число)
                span_p = driver.find_element(By.CSS_SELECTOR, 'span.ds-text.ds-text_weight_reg.ds-text_color_text-secondary.ds-text_proportional.ds-text_typography_lead-text.ds-text_lead-text_tight.ds-text_lead-text_reg')
                number = ''.join(filter(str.isdigit, span_p.text))  # Извлечение только цифр
                ws.cell(row=row, column=16, value=number)  # Колонка P
                
                print(f"Строка {row}: успешно обработана")
            except Exception as e:
                print(f"Строка {row}: не удалось извлечь данные со страницы - {str(e)}")
                
        except WebDriverException as e:
            print(f"Строка {row}: ошибка при загрузке страницы {link} - {str(e)}")
    
    # Сохранение изменений в Excel
    try:
        wb.save('index.xlsx')
        print("Изменения успешно сохранены в файл")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
    
    # Закрытие браузера
    driver.quit()

if __name__ == "__main__":
    main()