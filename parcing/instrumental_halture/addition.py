import openpyxl

def process_excel_file():
    # Загружаем исходный файл
    try:
        wb = openpyxl.load_workbook('index.xlsx')
        ws = wb.active
    except Exception as e:
        print(f"Ошибка при загрузке файла: {e}")
        return

    # Создаем новый файл для результатов
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active

    # Копируем все данные из исходного файла
    for row in ws.iter_rows():
        output_ws.append([cell.value for cell in row])

    # Обрабатываем данные согласно условиям
    for row in range(2, output_ws.max_row + 1):  # Начинаем с 2 строки, пропуская заголовок
        link_cell = output_ws.cell(row=row, column=14)  # Колонка N (14-я колонка)
        
        # Пропускаем строки без ссылки в колонке N
        if not link_cell.value or not isinstance(link_cell.value, str) or not link_cell.value.startswith(('http://', 'https://')):
            continue

        # Обработка колонки O
        o_cell = output_ws.cell(row=row, column=15)
        if o_cell.value is None or o_cell.value == "":
            o_cell.value = "-"
        elif isinstance(o_cell.value, (int, float)):
            if o_cell.value == int(o_cell.value):  # Проверяем, целое ли число
                o_cell.value = "-"

        # Обработка колонки P
        p_cell = output_ws.cell(row=row, column=16)
        if p_cell.value is None or p_cell.value == "":
            p_cell.value = "-"
        elif p_cell.value == 1 or p_cell.value == "1":  # Заменяем единицы на прочерк
            p_cell.value = "-"

    # Сохраняем результат в новый файл
    try:
        output_wb.save('output.xlsx')
        print("Файл успешно обработан и сохранен как output.xlsx")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")

if __name__ == "__main__":
    process_excel_file()