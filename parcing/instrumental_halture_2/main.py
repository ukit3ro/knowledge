import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import time
from webdriver_manager.chrome import ChromeDriverManager

def setup_driver():
    chrome_options = Options()
#    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-webgl")
    chrome_options.add_argument("--disable-3d-apis")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def parse_number(text):
    """Извлекает число из текста"""
    match = re.search(r'\d+', text.replace(' ', ''))
    return int(match.group()) if match else 0

def parse_reviews_with_selenium():
    try:
        # Загрузка Excel-файла
        wb = openpyxl.load_workbook('input.xlsx')
        sheet = wb.active
        print("Excel-файл успешно загружен")
    except Exception as e:
        print(f"Ошибка при загрузке Excel-файла: {e}")
        return

    driver = setup_driver()
    wait = WebDriverWait(driver, 10)

    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row=row, column=5).value  # Колонка E
        
        if not url or not isinstance(url, str) or not url.startswith(('http://', 'https://')):
            print(f"Строка {row}: Некорректный URL - {url}")
            continue

        try:
            print(f"Обработка строки {row}: {url}")
            driver.get(url)
            time.sleep(2)
            
            # Словарь для хранения количества отзывов по рейтингам
            ratings = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
            
            # Парсим данные для каждого рейтинга (1-5 звёзд)
            for stars in range(1, 6):
                try:
                    # Находим элемент с нужным рейтингом
                    element = wait.until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, f'a._9o0Y4g.BdsU1N.base-link input[name="rating"][value="{stars}"]')))
                    parent = element.find_element(By.XPATH, './ancestor::a[1]')
                    
                    # Извлекаем количество отзывов
                    rating_p = parent.find_element(
                        By.CSS_SELECTOR, 'p._typography_snzga_46._text_snzga_53._v2_snzga_15._-no-margin_snzga_49')
                    count = parse_number(rating_p.text)
                    ratings[stars] = count
                    
                    # Записываем в соответствующую колонку (I-M)
                    sheet.cell(row=row, column=8 + stars, value=count)
                except Exception as e:
                    print(f"Строка {row}: Ошибка при парсинге рейтинга {stars} - {str(e)}")
                    sheet.cell(row=row, column=8 + stars, value=0)

            # Получаем значения для каждого рейтинга
            i, j, k, l, m = ratings[1], ratings[2], ratings[3], ratings[4], ratings[5]
            
            # Вычисляем сумму всех отзывов (H)
            value_h = i + j + k + l + m
            sheet.cell(row=row, column=8, value=value_h)  # Колонка H
            
            # Вычисляем значение для колонки G: (i + j*2 + k*3 + l*4 + m*5) / H
            if value_h > 0:
                value_g = (i + j*2 + k*3 + l*4 + m*5) / value_h
            else:
                value_g = 0
                
            sheet.cell(row=row, column=7, value=round(value_g, 4))  # Колонка G
            
            print(f"Строка {row}: I={i} J={j} K={k} L={l} M={m} | G={round(value_g, 4)} H={value_h}")
            
            # Сохраняем после каждой строки
            wb.save('output.xlsx')
            time.sleep(1)
            
        except WebDriverException as e:
            print(f"Строка {row}: Ошибка при загрузке страницы - {str(e)}")
        except Exception as e:
            print(f"Строка {row}: Неожиданная ошибка - {str(e)}")

    # Завершение работы
    try:
        driver.quit()
        wb.save('output.xlsx')
        print("Парсинг завершен, результаты сохранены в output.xlsx")
    except Exception as e:
        print(f"Ошибка при завершении работы: {str(e)}")

if __name__ == "__main__":
    parse_reviews_with_selenium()